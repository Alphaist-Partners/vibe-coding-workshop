"""
exporter.py

Generates expense_report.xlsx (and pending_review.xlsx) based on learned_formats.json.
Renames and archives invoice files based on learned_naming.json.
"""

import json
import logging
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Excel generation
# ─────────────────────────────────────────────────────────────────────────────

# Default column structure for the report when no sample is available
DEFAULT_COLUMNS = [
    ("序号", "seq", 6),
    ("原始文件名", "original_filename", 25),
    ("新文件名", "new_filename", 30),
    ("文档类型", "doc_type", 18),
    ("日期", "date", 12),
    ("供应商", "vendor", 25),
    ("费用描述", "description", 30),
    ("金额", "amount", 12),
    ("货币", "currency", 8),
    ("税额", "tax_amount", 10),
    ("一级性质", "category_l1", 12),
    ("二级性质", "category_l2", 15),
    ("三级性质", "category_l3", 15),
    ("汇率", "exchange_rate", 10),
    ("人民币金额", "amount_cny", 14),
    ("发票代码", "invoice_code", 16),
    ("发票号码", "invoice_number", 16),
    ("信用卡后四位", "card_last_four", 12),
    ("分类置信度", "confidence", 10),
    ("备注", "notes", 30),
]


def _apply_cell_style(cell, style_dict: dict, wb=None) -> None:
    """Apply a style dict (from learned_formats) to an openpyxl cell."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if "font" in style_dict:
        fd = style_dict["font"]
        cell.font = Font(
            name=fd.get("name", "微软雅黑"),
            size=fd.get("size", 10),
            bold=fd.get("bold", False),
            italic=fd.get("italic", False),
            color=fd.get("color", "FF000000"),
        )

    if "fill" in style_dict:
        fd = style_dict["fill"]
        fg = fd.get("fgColor", "FFFFFFFF")
        pt = fd.get("patternType", "solid")
        if fg and fg not in ("FFFFFFFF", "00000000"):
            cell.fill = PatternFill(patternType=pt, fgColor=fg)

    if "alignment" in style_dict:
        ad = style_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=ad.get("horizontal", "general"),
            vertical=ad.get("vertical", "center"),
            wrap_text=ad.get("wrap_text", False),
        )

    if "border" in style_dict:
        bd = style_dict["border"]

        def _side(sd):
            if not sd:
                return Side()
            return Side(style=sd.get("style"), color=sd.get("color", "FF000000"))

        cell.border = Border(
            left=_side(bd.get("left")),
            right=_side(bd.get("right")),
            top=_side(bd.get("top")),
            bottom=_side(bd.get("bottom")),
        )

    if "number_format" in style_dict:
        cell.number_format = style_dict["number_format"]


def _build_default_workbook(records: list[dict], title: str = "费用报销明细"):
    """Build a default styled workbook when no sample format is available."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # Header styles
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFFFF")
    header_fill = PatternFill(patternType="solid", fgColor="FF336699")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, (header, _, width) in enumerate(DEFAULT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    # Data styles
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    center_align = Alignment(horizontal="center", vertical="center")
    number_align = Alignment(horizontal="right", vertical="center")
    data_font = Font(name="微软雅黑", size=9)
    alt_fill = PatternFill(patternType="solid", fgColor="FFF5F5F5")

    for row_idx, record in enumerate(records, 2):
        fill = alt_fill if row_idx % 2 == 0 else None

        for col_idx, (header, field, _) in enumerate(DEFAULT_COLUMNS, 1):
            if field == "seq":
                value = row_idx - 1
            else:
                value = record.get(field, "")
                if value is None:
                    value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            if fill:
                cell.fill = fill

            # Align numbers right
            if field in ("amount", "tax_amount", "amount_cny", "exchange_rate"):
                cell.alignment = number_align
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"
            elif field in ("seq",):
                cell.alignment = center_align
            else:
                cell.alignment = data_align

    # Summary row
    summary_row = len(records) + 2
    ws.cell(row=summary_row, column=1, value="合计").font = Font(name="微软雅黑", size=10, bold=True)

    # Find amount_cny column
    for col_idx, (_, field, _) in enumerate(DEFAULT_COLUMNS, 1):
        if field == "amount_cny":
            col_letter = get_column_letter(col_idx)
            total = sum(
                float(r.get("amount_cny", 0) or 0)
                for r in records
                if not r.get("skipped")
            )
            sum_cell = ws.cell(row=summary_row, column=col_idx, value=total)
            sum_cell.font = Font(name="微软雅黑", size=10, bold=True)
            sum_cell.number_format = "#,##0.00"
            sum_cell.alignment = number_align
            sum_cell.border = thin_border
            break

    return wb


def _build_workbook_from_format(
    records: list[dict],
    learned_formats: dict,
    title: str = "费用报销明细",
):
    """
    Build a workbook that replicates the sample format as closely as possible.
    Falls back to default if format info is incomplete.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sheets_info = learned_formats.get("sheets", {})
    if not sheets_info:
        return _build_default_workbook(records, title)

    # Use the first sheet as template
    sheet_name = list(sheets_info.keys())[0]
    sheet_fmt = sheets_info[sheet_name]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # Apply freeze panes
    freeze = sheet_fmt.get("freeze_panes")
    if freeze:
        try:
            ws.freeze_panes = freeze
        except Exception:
            ws.freeze_panes = "A2"

    # Get header rows from sample
    header_rows = sheet_fmt.get("header_rows", [])
    cell_styles = sheet_fmt.get("cell_styles", {})

    # Determine actual headers: last non-empty header row is the data column row
    data_header = []
    if header_rows:
        for row in reversed(header_rows):
            non_empty = [h for h in row if h]
            if non_empty:
                data_header = row
                break

    # Map sample headers to our data fields
    # Try to match by keywords
    field_mapping = _map_headers_to_fields(data_header)

    if not field_mapping:
        # Cannot map - fall back to default
        return _build_default_workbook(records, title)

    # Write header(s)
    for hr_idx, header_row_vals in enumerate(header_rows, 1):
        for col_idx, val in enumerate(header_row_vals, 1):
            if val:
                cell = ws.cell(row=hr_idx, column=col_idx, value=val)
                # Apply cell style from sample if available
                c_id = f"{get_column_letter(col_idx)}{hr_idx}"
                if c_id in cell_styles:
                    try:
                        _apply_cell_style(cell, cell_styles[c_id])
                    except Exception:
                        pass

    # Handle merged cells in header
    merged = sheet_fmt.get("merged_cells", [])
    for merge_range in merged:
        try:
            ws.merge_cells(merge_range)
        except Exception:
            pass

    # Column widths
    col_widths = sheet_fmt.get("column_widths", {})
    for col_letter, width in col_widths.items():
        try:
            ws.column_dimensions[col_letter].width = width
        except Exception:
            pass

    # Row heights
    row_heights = sheet_fmt.get("row_heights", {})
    for row_idx, height in row_heights.items():
        try:
            ws.row_dimensions[int(row_idx)].height = height
        except Exception:
            pass

    # Write data rows
    data_start_row = len(header_rows) + 1
    for rec_idx, record in enumerate(records, 0):
        row_idx = data_start_row + rec_idx
        for col_idx, field in sorted(field_mapping.items()):
            if field == "seq":
                value = rec_idx + 1
            else:
                value = record.get(field, "")
                if value is None:
                    value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Try to find a data row style in the sample (row data_start_row style)
            c_id = f"{get_column_letter(col_idx)}{data_start_row}"
            if c_id in cell_styles:
                try:
                    _apply_cell_style(cell, cell_styles[c_id])
                except Exception:
                    pass

    return wb


def _map_headers_to_fields(headers: list) -> dict[int, str]:
    """
    Map column indices (1-based) to data field names based on header keywords.
    """
    mapping: dict[int, str] = {}
    if not headers:
        return mapping

    # Keyword mappings: header keyword -> field name
    keyword_field_map = [
        (["序号", "no.", "#"], "seq"),
        (["原始文件", "文件名"], "original_filename"),
        (["新文件", "重命名"], "new_filename"),
        (["文档类型", "单据类型", "发票类型"], "doc_type"),
        (["日期", "开票日期", "消费日期", "date"], "date"),
        (["供应商", "卖方", "商户", "vendor"], "vendor"),
        (["描述", "摘要", "费用描述", "说明", "description"], "description"),
        (["金额", "价税合计", "消费金额", "amount"], "amount"),
        (["货币", "币种", "currency"], "currency"),
        (["税额", "税款", "tax"], "tax_amount"),
        (["一级", "category_l1", "费用类型"], "category_l1"),
        (["二级", "category_l2", "费用性质", "性质"], "category_l2"),
        (["三级", "category_l3"], "category_l3"),
        (["汇率", "exchange_rate", "rate"], "exchange_rate"),
        (["人民币", "cny", "换算"], "amount_cny"),
        (["发票代码", "invoice_code"], "invoice_code"),
        (["发票号", "invoice_number", "号码"], "invoice_number"),
        (["卡号", "信用卡", "card"], "card_last_four"),
        (["置信度", "confidence"], "confidence"),
        (["备注", "notes", "remark"], "notes"),
    ]

    for col_idx, header_val in enumerate(headers, 1):
        if not header_val:
            continue
        header_lower = str(header_val).lower().strip()
        for keywords, field in keyword_field_map:
            if any(kw in header_lower for kw in keywords):
                if col_idx not in mapping:  # first match wins
                    mapping[col_idx] = field
                break

    return mapping


def export_report(
    records: list[dict],
    output_path: str,
    learned_formats: dict,
    title: str = "费用报销明细",
) -> str:
    """
    Generate expense_report.xlsx from records.

    Returns the output path.
    """
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    confidence = learned_formats.get("confidence", "low")
    if confidence in ("high", "medium") and learned_formats.get("sheets"):
        wb = _build_workbook_from_format(records, learned_formats, title)
    else:
        wb = _build_default_workbook(records, title)

    wb.save(str(out_path))
    logger.info(f"Saved expense report to {out_path}")
    return str(out_path)


def export_pending_review(
    pending_records: list[dict],
    output_path: str,
) -> Optional[str]:
    """
    Generate pending_review.xlsx for records needing manual review.

    Returns output path or None if no pending records.
    """
    if not pending_records:
        return None

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = _build_default_workbook(pending_records, "待确认报销明细")

    # Add a "待确认原因" column note
    ws = wb.active
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.cell(row=1, column=len(DEFAULT_COLUMNS) + 1, value="待确认原因").font = Font(
        name="微软雅黑", size=10, bold=True, color="FFFFFFFF"
    )
    ws.cell(row=1, column=len(DEFAULT_COLUMNS) + 1).fill = PatternFill(
        patternType="solid", fgColor="FFCC3300"
    )

    for row_idx, record in enumerate(pending_records, 2):
        reason = record.get("pending_reason", "需要人工确认")
        ws.cell(row=row_idx, column=len(DEFAULT_COLUMNS) + 1, value=reason)

    wb.save(str(out_path))
    logger.info(f"Saved pending review to {out_path}")
    return str(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# File renaming and archiving
# ─────────────────────────────────────────────────────────────────────────────

def rename_and_archive(
    records: list[dict],
    output_renamed_base: str,
    person: str,
) -> dict[str, str]:
    """
    Copy/rename invoice files to output/renamed_files/{person}/ directory.

    Updates each record's 'archived_path' field.
    Returns dict mapping original_filename -> new_path.
    """
    person_dir = Path(output_renamed_base) / person
    person_dir.mkdir(parents=True, exist_ok=True)

    rename_map: dict[str, str] = {}
    seq_counter: dict[str, int] = {}  # track sequence per (date, category) combo

    for record in records:
        if record.get("skipped"):
            continue

        source_path = record.get("source_path", "")
        new_filename = record.get("new_filename", "")
        original_filename = record.get("original_filename", "")

        if not source_path or not Path(source_path).exists():
            logger.warning(f"Source file not found: {source_path}")
            continue

        if not new_filename:
            new_filename = original_filename

        # Ensure no path traversal
        new_filename = Path(new_filename).name

        dest_path = person_dir / new_filename

        # Handle name collisions
        if dest_path.exists():
            stem = dest_path.stem
            suffix = dest_path.suffix
            counter = 1
            while dest_path.exists():
                dest_path = person_dir / f"{stem}_{counter}{suffix}"
                counter += 1

        try:
            shutil.copy2(source_path, dest_path)
            record["archived_path"] = str(dest_path)
            rename_map[original_filename] = str(dest_path)
            logger.info(f"Archived: {original_filename} -> {dest_path.name}")
        except Exception as e:
            logger.error(f"Failed to archive {source_path}: {e}")
            record["archived_path"] = ""

    return rename_map


def generate_new_filenames(
    records: list[dict],
    learned_naming: dict,
    person: str,
) -> None:
    """
    Populate 'new_filename' field in each record based on learned naming convention.
    Modifies records in place.
    """
    from src.learn_naming import apply_naming_pattern

    # Track sequence numbers per (date, category) combo
    seq_tracker: dict[str, int] = {}

    for record in records:
        if record.get("skipped"):
            continue

        date_val = record.get("date", "")
        category_l2 = record.get("category_l2", "其他")
        amount = record.get("amount", "0")
        currency = record.get("currency", "CNY")
        doc_type = record.get("doc_type", "")
        original_filename = record.get("original_filename", "")
        original_ext = Path(original_filename).suffix if original_filename else ".pdf"

        # Sequence key
        seq_key = f"{date_val}_{category_l2}"
        seq_tracker[seq_key] = seq_tracker.get(seq_key, 0) + 1
        seq = seq_tracker[seq_key]

        new_name = apply_naming_pattern(
            naming=learned_naming,
            person=person,
            date=date_val,
            category_l2=category_l2,
            amount=amount,
            currency=currency,
            seq=seq,
            original_ext=original_ext,
            doc_type=doc_type,
        )

        record["new_filename"] = new_name
