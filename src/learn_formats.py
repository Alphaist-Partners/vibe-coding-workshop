"""
learn_formats.py

Reads reference/sample_report.xlsx and learns:
  - Excel report format (headers, styles, widths, colors, merged cells, etc.)
  - Expense category rules from content

Outputs:
  - reference/learned_formats.json
  - reference/category_rules.json
"""

import json
import logging
import os
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def _safe_color(color_obj) -> str | None:
    """Extract hex color string safely from openpyxl color objects."""
    if color_obj is None:
        return None
    try:
        t = getattr(color_obj, "type", None)
        if t == "rgb":
            v = getattr(color_obj, "rgb", None)
            if v and v not in ("00000000", "FF000000", "FFFFFFFF"):
                return str(v)
        elif t == "theme":
            return f"theme:{color_obj.theme}"
    except Exception:
        pass
    return None


def _font_to_dict(font) -> dict:
    if font is None:
        return {}
    d: dict[str, Any] = {}
    if getattr(font, "name", None):
        d["name"] = font.name
    if getattr(font, "size", None):
        d["size"] = font.size
    if getattr(font, "bold", None):
        d["bold"] = font.bold
    if getattr(font, "italic", None):
        d["italic"] = font.italic
    c = _safe_color(getattr(font, "color", None))
    if c:
        d["color"] = c
    return d


def _fill_to_dict(fill) -> dict:
    if fill is None:
        return {}
    d: dict[str, Any] = {}
    fg = _safe_color(getattr(fill, "fgColor", None))
    if fg:
        d["fgColor"] = fg
    pt = getattr(fill, "patternType", None)
    if pt:
        d["patternType"] = str(pt)
    return d


def _alignment_to_dict(alignment) -> dict:
    if alignment is None:
        return {}
    d: dict[str, Any] = {}
    for attr in ("horizontal", "vertical", "wrap_text", "shrink_to_fit"):
        val = getattr(alignment, attr, None)
        if val is not None:
            d[attr] = val
    return d


def _border_side_to_dict(side) -> dict:
    if side is None:
        return {}
    d: dict[str, Any] = {}
    if getattr(side, "style", None):
        d["style"] = side.style
    c = _safe_color(getattr(side, "color", None))
    if c:
        d["color"] = c
    return d


def _border_to_dict(border) -> dict:
    if border is None:
        return {}
    d: dict[str, Any] = {}
    for side in ("left", "right", "top", "bottom"):
        s = getattr(border, side, None)
        sd = _border_side_to_dict(s)
        if sd:
            d[side] = sd
    return d


def _extract_sheet_formats(ws) -> dict:
    """Extract comprehensive formatting info from a worksheet."""
    fmt: dict[str, Any] = {}

    # Sheet dimensions
    fmt["max_row"] = ws.max_row
    fmt["max_column"] = ws.max_column

    # Column widths
    col_widths = {}
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.width:
            col_widths[col_letter] = col_dim.width
    fmt["column_widths"] = col_widths

    # Row heights
    row_heights = {}
    for row_idx, row_dim in ws.row_dimensions.items():
        if row_dim.height:
            row_heights[row_idx] = row_dim.height
    fmt["row_heights"] = row_heights

    # Merged cells
    merged = []
    for merge_range in ws.merged_cells.ranges:
        merged.append(str(merge_range))
    fmt["merged_cells"] = merged

    # Freeze panes
    if ws.freeze_panes:
        fmt["freeze_panes"] = str(ws.freeze_panes)

    # Tab color
    if ws.sheet_properties and ws.sheet_properties.tabColor:
        c = _safe_color(ws.sheet_properties.tabColor)
        if c:
            fmt["tab_color"] = c

    # Sheet view
    fmt["show_gridlines"] = True  # default
    if ws.sheet_view:
        fmt["show_gridlines"] = ws.sheet_view.showGridLines

    # Per-cell styles (first 50 rows to capture headers + sample data)
    cell_styles = {}
    for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row)):
        for cell in row:
            c_id = f"{cell.column_letter}{cell.row}"
            style: dict[str, Any] = {}
            font_d = _font_to_dict(cell.font)
            if font_d:
                style["font"] = font_d
            fill_d = _fill_to_dict(cell.fill)
            if fill_d:
                style["fill"] = fill_d
            align_d = _alignment_to_dict(cell.alignment)
            if align_d:
                style["alignment"] = align_d
            border_d = _border_to_dict(cell.border)
            if border_d:
                style["border"] = border_d
            if cell.number_format and cell.number_format != "General":
                style["number_format"] = cell.number_format
            if style:
                cell_styles[c_id] = style
    fmt["cell_styles"] = cell_styles

    # Headers (row 1 and row 2 if merged header rows)
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        row_vals = [str(v) if v is not None else "" for v in row]
        if any(row_vals):
            headers.append(row_vals)
    fmt["header_rows"] = headers

    # Conditional formatting summary
    cond_fmts = []
    for sqref, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            cond_fmts.append({
                "sqref": str(sqref),
                "type": getattr(rule, "type", None),
                "priority": getattr(rule, "priority", None),
            })
    fmt["conditional_formatting"] = cond_fmts

    return fmt


def _extract_category_rules(ws) -> list[dict]:
    """
    Scan sheet data rows and extract category rules.
    Looks for columns that seem to be expense category columns (l1, l2, l3).
    """
    rules = []
    if ws.max_row < 2:
        return rules

    # Find header row
    header_row = None
    header_idx = 1
    for row_idx in range(1, min(5, ws.max_row + 1)):
        row_vals = [cell.value for cell in ws[row_idx]]
        non_none = [v for v in row_vals if v is not None]
        if len(non_none) >= 3:
            header_row = row_vals
            header_idx = row_idx
            break

    if not header_row:
        return rules

    # Map column index -> header name
    col_map = {i: str(v).strip() if v else "" for i, v in enumerate(header_row)}

    # Keywords to identify category columns
    cat_keywords = {
        "l1": ["一级", "category_l1", "费用性质", "费用类型", "大类"],
        "l2": ["二级", "category_l2", "费用细类", "性质", "小类"],
        "l3": ["三级", "category_l3"],
        "vendor": ["供应商", "vendor", "卖方", "商户"],
        "description": ["描述", "摘要", "说明", "description", "备注"],
        "amount": ["金额", "amount", "价税合计"],
    }

    col_roles: dict[str, int] = {}
    for i, header in col_map.items():
        h_lower = header.lower()
        for role, keywords in cat_keywords.items():
            if any(kw in h_lower for kw in keywords):
                col_roles[role] = i
                break

    # Extract unique category combinations from data rows
    seen = set()
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        entry: dict[str, str] = {}
        for role, col_idx in col_roles.items():
            if col_idx < len(row) and row[col_idx] is not None:
                entry[role] = str(row[col_idx]).strip()

        if not entry:
            continue

        key = json.dumps(entry, sort_keys=True, ensure_ascii=False)
        if key not in seen:
            seen.add(key)
            rules.append(entry)

    return rules


def _build_default_category_rules() -> list[dict]:
    """Return sensible default category rules for a GP fund."""
    return [
        {"l1": "差旅费", "l2": "机票", "l3": "", "keywords": ["机票", "航空", "airline", "flight"]},
        {"l1": "差旅费", "l2": "火车票", "l3": "", "keywords": ["火车", "高铁", "铁路", "railway"]},
        {"l1": "差旅费", "l2": "酒店", "l3": "", "keywords": ["酒店", "hotel", "住宿", "旅馆"]},
        {"l1": "差旅费", "l2": "出租车/网约车", "l3": "", "keywords": ["出租", "网约车", "滴滴", "uber", "taxi", "grab"]},
        {"l1": "差旅费", "l2": "地铁/公交", "l3": "", "keywords": ["地铁", "公交", "地铁票"]},
        {"l1": "餐饮费", "l2": "工作餐", "l3": "", "keywords": ["餐饮", "餐厅", "restaurant", "food", "饭店", "咖啡"]},
        {"l1": "餐饮费", "l2": "接待餐", "l3": "", "keywords": ["接待", "宴请"]},
        {"l1": "办公费", "l2": "办公用品", "l3": "", "keywords": ["文具", "办公", "打印", "耗材"]},
        {"l1": "办公费", "l2": "软件订阅", "l3": "", "keywords": ["软件", "订阅", "saas", "cloud", "subscription"]},
        {"l1": "通讯费", "l2": "电话费", "l3": "", "keywords": ["电话", "手机", "通讯", "话费"]},
        {"l1": "通讯费", "l2": "网络费", "l3": "", "keywords": ["宽带", "网络", "wifi", "internet"]},
        {"l1": "专业服务费", "l2": "法务", "l3": "", "keywords": ["律师", "法律", "legal", "law"]},
        {"l1": "专业服务费", "l2": "审计/财务", "l3": "", "keywords": ["审计", "会计", "财务", "audit"]},
        {"l1": "专业服务费", "l2": "咨询", "l3": "", "keywords": ["咨询", "consulting", "顾问"]},
        {"l1": "会务费", "l2": "会议", "l3": "", "keywords": ["会议", "conference", "论坛", "forum"]},
        {"l1": "其他", "l2": "其他", "l3": "", "keywords": []},
    ]


def learn_formats(config: dict, force: bool = False) -> tuple[dict, list[dict]]:
    """
    Main entry: read sample_report.xlsx, produce learned_formats and category_rules.

    Returns (learned_formats_dict, category_rules_list)
    """
    ref = config.get("reference", {})
    sample_path = Path(ref.get("sample_report", "./reference/sample_report.xlsx"))
    learned_formats_path = Path(ref.get("learned_formats", "./reference/learned_formats.json"))
    category_rules_path = Path(ref.get("category_rules", "./reference/category_rules.json"))

    # Check if already learned and not forced
    if not force and learned_formats_path.exists() and category_rules_path.exists():
        logger.info("Learned formats already exist. Loading from cache.")
        with open(learned_formats_path, encoding="utf-8") as f:
            learned = json.load(f)
        with open(category_rules_path, encoding="utf-8") as f:
            rules = json.load(f)
        return learned, rules

    learned_formats: dict[str, Any] = {
        "source": str(sample_path),
        "sheets": {},
        "confidence": "medium",
    }
    category_rules: list[dict] = []

    if not sample_path.exists():
        logger.warning(f"sample_report.xlsx not found at {sample_path}. Using defaults.")
        learned_formats["confidence"] = "low"
        learned_formats["note"] = "No sample file found. Using default format."
        category_rules = _build_default_category_rules()
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(sample_path), data_only=True)
            learned_formats["sheet_names"] = wb.sheetnames
            learned_formats["confidence"] = "high"

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"Analyzing sheet: {sheet_name}")
                sheet_fmt = _extract_sheet_formats(ws)
                learned_formats["sheets"][sheet_name] = sheet_fmt

                # Try extracting categories from every sheet
                cats = _extract_category_rules(ws)
                for c in cats:
                    # Avoid exact duplicates
                    if c not in category_rules:
                        category_rules.append(c)

            if not category_rules:
                logger.info("No category data found in sample. Using defaults.")
                category_rules = _build_default_category_rules()

            # Record workbook-level properties
            learned_formats["named_styles"] = [s.name for s in wb.named_styles] if wb.named_styles else []

        except Exception as e:
            logger.error(f"Error reading sample_report.xlsx: {e}")
            learned_formats["confidence"] = "low"
            learned_formats["error"] = str(e)
            category_rules = _build_default_category_rules()

    # Persist
    learned_formats_path.parent.mkdir(parents=True, exist_ok=True)
    with open(learned_formats_path, "w", encoding="utf-8") as f:
        json.dump(learned_formats, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved learned_formats.json to {learned_formats_path}")

    with open(category_rules_path, "w", encoding="utf-8") as f:
        json.dump(category_rules, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved category_rules.json to {category_rules_path}")

    return learned_formats, category_rules


def summarize_formats(learned_formats: dict, category_rules: list[dict]) -> str:
    """Return a human-readable summary for terminal display."""
    lines = []
    confidence = learned_formats.get("confidence", "unknown")
    lines.append(f"[Step 0a] Excel 格式学习结果 (置信度: {confidence})")

    sheets = learned_formats.get("sheets", {})
    if sheets:
        for sname, sfmt in sheets.items():
            lines.append(f"  Sheet: {sname}")
            headers = sfmt.get("header_rows", [])
            if headers:
                lines.append(f"    表头行数: {len(headers)}")
                lines.append(f"    第一行表头: {[h for h in headers[0] if h]}")
            col_widths = sfmt.get("column_widths", {})
            lines.append(f"    列数: {sfmt.get('max_column', '?')}")
            lines.append(f"    自定义列宽数: {len(col_widths)}")
            merged = sfmt.get("merged_cells", [])
            lines.append(f"    合并单元格数: {len(merged)}")
            freeze = sfmt.get("freeze_panes", None)
            if freeze:
                lines.append(f"    冻结窗格: {freeze}")
    else:
        lines.append("  未找到样本文件，将使用默认格式")

    lines.append(f"[Step 0a] 分类规则学习结果: 共 {len(category_rules)} 条规则")
    unique_l1 = list({r.get("l1", r.get("category_l1", "")) for r in category_rules if r.get("l1") or r.get("category_l1")})
    if unique_l1:
        lines.append(f"  一级分类: {unique_l1}")

    return "\n".join(lines)
